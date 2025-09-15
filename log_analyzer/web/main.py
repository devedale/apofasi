import sys
import os
import json
import csv
import glob
import shutil
import asyncio
import time
import uuid
from queue import Queue, Empty
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
from threading import Thread


from fastapi import FastAPI, Request, UploadFile, File, Form
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel

# --- Path Setup ---
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))

# --- Service-based Imports ---
from log_analyzer.services.config_service import ConfigService
from log_analyzer.services.logppt_service import LogPPTService
from log_analyzer.services.logppt2_service import LogPPT2Service
from log_analyzer.services.presidio_service import PresidioService
from log_analyzer.services.ollama_service import OllamaService
from log_analyzer.services.llm_finetuning_service import LLMFineTuningService
from log_analyzer.parsing.interfaces import LogEntry, ParsedRecord
from log_analyzer.parsing.parser_factory import create_parser_chain
from log_analyzer.services.log_reader import LogReader
from log_analyzer.services.drain3_service import Drain3Service
from presidio_analyzer import AnalyzerEngine, RecognizerRegistry, Pattern, PatternRecognizer
from presidio_anonymizer import AnonymizerEngine, OperatorConfig


# --- FastAPI App Initialization ---
import logging

# Reduce noisy logs from presidio-analyzer and its submodules
logging.getLogger("presidio").setLevel(logging.WARNING)
logging.getLogger("presidio-analyzer").setLevel(logging.WARNING)
logging.getLogger("presidio_analyzer").setLevel(logging.WARNING)
logging.getLogger("presidio_anonymizer").setLevel(logging.WARNING)

app = FastAPI()
app.mount("/static", StaticFiles(directory="log_analyzer/web/static"), name="static")
app.mount("/outputs", StaticFiles(directory="outputs"), name="outputs")
templates = Jinja2Templates(directory="log_analyzer/web/templates")

# --- Pydantic Models ---
class PreviewRequest(BaseModel):
    sample_text: str
    presidio_config: Dict[str, Any]

class ConfigUpdateRequest(BaseModel):
    presidio: Dict[str, Any]

class AnalysisRequest(BaseModel):
    input_file: str

class ModelDownloadRequest(BaseModel):
    model_name: str

# --- Fine-Tuning Models ---
class DatasetValidationRequest(BaseModel):
    dataset_path: str
    format_type: str

class FineTuningRequest(BaseModel):
    base_model: str
    dataset_path: str
    model_name: str
    training_template: str
    custom_config: Optional[Dict[str, Any]] = None

class ModelTestRequest(BaseModel):
    model_name: str
    test_prompt: str

# --- Real-time Logging for Backend Tasks ---
class TaskLogManager:
    def __init__(self):
        self.tasks = {}

    def start_task(self):
        task_id = str(uuid.uuid4())
        self.tasks[task_id] = Queue()
        return task_id

    def log(self, task_id, message):
        if task_id in self.tasks:
            self.tasks[task_id].put(message)

    def finish_task(self, task_id):
        if task_id in self.tasks:
            self.tasks[task_id].put(None) # Sentinel value to indicate end of stream

    def get_log_stream(self, task_id):
        if task_id not in self.tasks:
            return

        queue = self.tasks[task_id]
        while True:
            try:
                message = queue.get_nowait()
                if message is None:
                    break
                yield f"data: {message}\n\n"
            except Empty:
                yield "" # Keep connection alive
                time.sleep(0.1)

task_log_manager = TaskLogManager()

@app.get("/api/stream-logs/{task_id}")
async def stream_logs(task_id: str):
    return StreamingResponse(task_log_manager.get_log_stream(task_id), media_type="text/event-stream")


# --- Model Management (now handled by Ollama) ---
# MODELS_PATH removed - models are now managed by Ollama service

def get_dir_size(path='.'):
    total_size = 0
    for dirpath, dirnames, filenames in os.walk(path):
        for f in filenames:
            fp = os.path.join(dirpath, f)
            # skip if it is symbolic link
            if not os.path.islink(fp):
                total_size += os.path.getsize(fp)
    return total_size

@app.get("/api/models", response_class=JSONResponse)
async def get_models():
    """
    Lista tutti i modelli disponibili in Ollama.
    """
    try:
        config_service = ConfigService()
        config = config_service.load_config()
        
        # Crea istanza del servizio Ollama
        ollama_service = OllamaService(config)
        
        # Ottieni i modelli da Ollama
        models = await ollama_service.list_models()
        
        # Ottieni informazioni sui modelli predefiniti
        available_models_info = ollama_service.get_available_models_info()
        
        # Combina le informazioni
        models_with_info = []
        for model in models:
            model_name = model["name"].split(":")[0]  # Rimuovi tag
            model_info = {
                "name": model["name"],
                "size_mb": round(model.get("size", 0) / (1024 * 1024), 2),
                "modified_at": model.get("modified_at", ""),
                "description": model.get("description", "")
            }
            
            # Aggiungi descrizione se è un modello predefinito
            if model_name in available_models_info:
                model_info["description"] = available_models_info[model_name]["description"]
                model_info["is_predefined"] = True
            else:
                model_info["is_predefined"] = False
            
            models_with_info.append(model_info)
        
        return {
            "models": sorted(models_with_info, key=lambda x: x['name']),
            "available_models": available_models_info
        }
        
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"Failed to list models: {str(e)}"})

@app.delete("/api/models/{model_name}", response_class=JSONResponse)
async def delete_model(model_name: str):
    """
    Elimina un modello da Ollama.
    """
    try:
        config_service = ConfigService()
        config = config_service.load_config()
        
        # Crea istanza del servizio Ollama
        ollama_service = OllamaService(config)
        
        # Elimina il modello da Ollama
        result = await ollama_service.delete_model(model_name)
        
        if result.get("success", False):
            return {"message": result.get("message", f"Model '{model_name}' deleted successfully.")}
        else:
            return JSONResponse(status_code=400, content={"error": result.get("error", "Failed to delete model")})
            
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"An error occurred during model deletion: {str(e)}"})

@app.post("/api/models/download", response_class=JSONResponse)
async def download_model(request: ModelDownloadRequest):
    """
    Scarica/crea un modello in Ollama.
    """
    try:
        model_name = request.model_name
        if not model_name or not isinstance(model_name, str):
            return JSONResponse(status_code=400, content={"error": "Invalid model name provided."})

        task_id = task_log_manager.start_task()

        def do_download(task_id):
            task_log_manager.log(task_id, f"Starting download/creation for model: {model_name}")
            try:
                config_service = ConfigService()
                config = config_service.load_config()
                
                # Crea istanza del servizio Ollama
                ollama_service = OllamaService(config, task_id, task_log_manager)
                
                # Usa asyncio per eseguire il download
                import asyncio
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                
                result = loop.run_until_complete(ollama_service.download_model(model_name))
                
                if result.get("success", False):
                    task_log_manager.log(task_id, f"Successfully downloaded/created model: {model_name}")
                else:
                    task_log_manager.log(task_id, f"Failed to download model {model_name}: {result.get('error', 'Unknown error')}")
                    
                loop.close()
                
            except Exception as e:
                task_log_manager.log(task_id, f"An unexpected error occurred during model download for {model_name}: {e}")
            finally:
                task_log_manager.finish_task(task_id)

        thread = Thread(target=do_download, args=(task_id,))
        thread.start()

        return {"task_id": task_id}

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": f"An error occurred during model download: {str(e)}"})

@app.get("/api/ollama/health", response_class=JSONResponse)
async def ollama_health_check():
    """
    Verifica lo stato di salute del servizio Ollama.
    """
    try:
        config_service = ConfigService()
        config = config_service.load_config()
        
        ollama_service = OllamaService(config)
        health = await ollama_service.health_check()
        
        return {
            "status": "healthy" if health else "unhealthy",
            "available": health,
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        return JSONResponse(status_code=500, content={
            "status": "error",
            "available": False,
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        })

# --- LogPPT 2 Config & Datasets ---
@app.get("/api/logppt2/datasets", response_class=JSONResponse)
async def get_logppt2_datasets():
    try:
        config_service = ConfigService()
        config = config_service.load_config()
        lp2 = config.get('logppt2', {})
        dataset_dir = Path(lp2.get('dataset_dir', 'examples'))
        dataset_dir.mkdir(exist_ok=True)
        files = []
        try:
            for f in dataset_dir.iterdir():
                if f.is_file() and f.suffix.lower() == '.csv':
                    files.append(f.name)
        except Exception:
            files = []
        return {"dataset_dir": str(dataset_dir), "files": sorted(files)}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


class LogPPT2ConfigRequest(BaseModel):
    dataset_dir: str


@app.post("/api/logppt2/config", response_class=JSONResponse)
async def save_logppt2_config(req: LogPPT2ConfigRequest):
    try:
        path = Path(req.dataset_dir)
        if not path.exists() or not path.is_dir():
            return JSONResponse(status_code=400, content={"error": f"Directory non valida: {path}"})
        config_service = ConfigService()
        full_config = config_service.load_config()
        if 'logppt2' not in full_config:
            full_config['logppt2'] = {}
        full_config['logppt2']['dataset_dir'] = str(path)
        if config_service.save_config(full_config):
            return {"status": "success", "message": "Configurazione LogPPT 2 salvata."}
        return JSONResponse(status_code=500, content={"error": "Salvataggio configurazione fallito."})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# --- LLM Fine-Tuning API ---
@app.get("/api/finetuning/base-models", response_class=JSONResponse)
async def get_base_models():
    """
    Ottiene la lista dei modelli base disponibili per il fine-tuning.
    """
    try:
        config_service = ConfigService()
        config = config_service.load_config()
        
        finetuning_service = LLMFineTuningService(config)
        base_models = await finetuning_service.get_available_base_models()
        
        return {"base_models": base_models}
        
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"Errore nel recupero dei modelli base: {str(e)}"})

@app.get("/api/finetuning/templates", response_class=JSONResponse)
async def get_training_templates():
    """
    Ottiene tutti i template di training disponibili.
    """
    try:
        config_service = ConfigService()
        config = config_service.load_config()
        
        finetuning_service = LLMFineTuningService(config)
        templates = finetuning_service.get_training_templates()
        
        return {"templates": templates}
        
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"Errore nel recupero dei template: {str(e)}"})

@app.post("/api/finetuning/validate-dataset", response_class=JSONResponse)
async def validate_dataset(request: DatasetValidationRequest):
    """
    Valida un dataset per il fine-tuning.
    """
    try:
        config_service = ConfigService()
        config = config_service.load_config()
        
        finetuning_service = LLMFineTuningService(config)
        validation_result = await finetuning_service.validate_dataset(
            request.dataset_path, 
            request.format_type
        )
        
        return validation_result
        
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"Errore nella validazione: {str(e)}"})

@app.post("/api/finetuning/start", response_class=JSONResponse)
async def start_finetuning(request: FineTuningRequest):
    """
    Avvia il processo di fine-tuning.
    """
    try:
        task_id = task_log_manager.start_task()
        
        def do_finetuning(task_id):
            try:
                config_service = ConfigService()
                config = config_service.load_config()
                
                finetuning_service = LLMFineTuningService(config, task_id, task_log_manager)
                
                import asyncio
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                
                result = loop.run_until_complete(finetuning_service.start_finetuning(
                    base_model=request.base_model,
                    dataset_path=request.dataset_path,
                    model_name=request.model_name,
                    training_template=request.training_template,
                    custom_config=request.custom_config
                ))
                
                if result["success"]:
                    task_log_manager.log(task_id, f"Fine-tuning completato: {request.model_name}")
                else:
                    task_log_manager.log(task_id, f"Fine-tuning fallito: {result.get('error', 'Errore sconosciuto')}")
                
                loop.close()
                
            except Exception as e:
                task_log_manager.log(task_id, f"Errore durante il fine-tuning: {str(e)}")
            finally:
                task_log_manager.finish_task(task_id)
        
        thread = Thread(target=do_finetuning, args=(task_id,))
        thread.start()
        
        return {"task_id": task_id, "message": "Fine-tuning avviato"}
        
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"Errore nell'avvio del fine-tuning: {str(e)}"})

@app.get("/api/finetuning/models", response_class=JSONResponse)
async def get_finetuned_models():
    """
    Lista tutti i modelli fine-tuned disponibili.
    """
    try:
        config_service = ConfigService()
        config = config_service.load_config()
        
        finetuning_service = LLMFineTuningService(config)
        models = await finetuning_service.list_finetuned_models()
        
        return {"models": models}
        
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"Errore nel recupero dei modelli: {str(e)}"})

@app.post("/api/finetuning/test", response_class=JSONResponse)
async def test_model(request: ModelTestRequest):
    """
    Testa un modello fine-tuned con un prompt.
    """
    try:
        config_service = ConfigService()
        config = config_service.load_config()
        
        finetuning_service = LLMFineTuningService(config)
        test_result = await finetuning_service.test_model(
            request.model_name,
            request.test_prompt
        )
        
        return test_result
        
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"Errore nel test del modello: {str(e)}"})

@app.post("/api/finetuning/export/{model_name}", response_class=JSONResponse)
async def export_model(model_name: str):
    """
    Esporta un modello fine-tuned.
    """
    try:
        config_service = ConfigService()
        config = config_service.load_config()
        
        finetuning_service = LLMFineTuningService(config)
        export_result = await finetuning_service.export_model(model_name, f"exports/{model_name}")
        
        return export_result
        
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"Errore nell'export del modello: {str(e)}"})


# --- Formatting Helper Functions ---
def _build_operators(config: Dict[str, Any]) -> Dict[str, OperatorConfig]:
    """Helper to build the operators dictionary for the AnonymizerEngine."""
    anonymizer_config = config.get("anonymizer", {})
    strategies = anonymizer_config.get("strategies", {})
    strategy_configs = anonymizer_config.get("strategy_config", {})

    operators = {}
    for entity_name, strategy_name in strategies.items():
        if strategy_name in strategy_configs:
            # If a detailed configuration exists for this strategy, use it
            params = strategy_configs[strategy_name].copy()  # Copy to avoid modifying original
            
            # Fix parameter mapping for mask strategy
            if strategy_name == "mask":
                # Map mask_char to masking_char
                if "mask_char" in params:
                    params["masking_char"] = params.pop("mask_char")
                
                # Add missing required parameters for mask operator
                if "chars_to_mask" not in params:
                    params["chars_to_mask"] = 0  # Default: mask all characters
                
                if "masking_char" not in params:
                    params["masking_char"] = "*"  # Default masking character
                
                if "from_end" not in params:
                    params["from_end"] = False  # Default: mask from beginning
            
            operators[entity_name] = OperatorConfig(strategy_name, params)
        else:
            # Otherwise, use the operator with its default parameters
            operators[entity_name] = OperatorConfig(strategy_name)

    # Also handle ad-hoc recognizers, which might have their own strategy defined
    ad_hoc_recognizers = config.get('analyzer', {}).get('ad_hoc_recognizers', [])
    for rec_conf in ad_hoc_recognizers:
        strategy_name = rec_conf.get("strategy")
        entity_name = rec_conf.get("name")
        if entity_name and strategy_name:
             if strategy_name in strategy_configs:
                params = strategy_configs[strategy_name].copy()  # Copy to avoid modifying original
                
                # Fix parameter mapping for mask strategy
                if strategy_name == "mask":
                    # Map mask_char to masking_char
                    if "mask_char" in params:
                        params["masking_char"] = params.pop("mask_char")
                    
                    # Add missing required parameters for mask operator
                    if "chars_to_mask" not in params:
                        params["chars_to_mask"] = 0  # Default: mask all characters
                    
                    if "masking_char" not in params:
                        params["masking_char"] = "*"  # Default masking character
                    
                    if "from_end" not in params:
                        params["from_end"] = False  # Default: mask from beginning
                
                operators[entity_name] = OperatorConfig(strategy_name, params)
             else:
                operators[entity_name] = OperatorConfig(strategy_name)

    return operators

def format_as_anonymized_text(records: List[ParsedRecord], output_path: str):
    with open(output_path, "w", encoding="utf-8") as f:
        for record in records:
            f.write(f"{record.presidio_anonymized or ''}\n")

def format_as_logppt(records: List[ParsedRecord], output_path: str, version: str = "anonymized"):
    """
    Generate LogPPT CSV report for either original or anonymized data.
    
    Args:
        records: List of parsed records
        output_path: Output file path
        version: "original" or "anonymized" to determine which data to use
    """
    if not records: return
    
    all_keys = set()
    for record in records:
        if record.parsed_data: all_keys.update(record.parsed_data.keys())
    sorted_keys = sorted(list(all_keys))
    headers = ["LineId", "Timestamp"] + sorted_keys + ["Content", "EventId", "Template"]
    
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        
        for record in records:
            # Choose drain3 results based on version
            if version == "original":
                drain_result = record.drain3_original or {}
                content = record.original_content
            else:  # anonymized
                drain_result = record.drain3_anonymized or {}
                content = record.presidio_anonymized or record.original_content
            
            # Extract timestamp from parsed data, fallback to "N/A" if not found
            timestamp = record.parsed_data.get("timestamp", "N/A")
            # Format EventId as "E{cluster_id}" with fallback to "N/A"
            event_id = f"E{drain_result.get('cluster_id', 'N/A')}" if drain_result.get('cluster_id') is not None else "N/A"
            # Get template from drain3 result, fallback to "N/A"
            template = drain_result.get('template', 'N/A')
            
            row = {"LineId": record.line_number, "Timestamp": timestamp, "Content": content, "EventId": event_id, "Template": template}
            for key in sorted_keys:
                row[key] = record.parsed_data.get(key, "")
            writer.writerow(row)

def format_as_json_report(records: List[ParsedRecord], output_path: str):
    report = [record.model_dump(exclude_none=True) for record in records]
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2)

def format_as_csv_from_unified(unified_data: List[Dict], output_path: str, version: str = "anonymized"):
    """
    Generate clean CSV export from unified log data.
    This function ensures CSV compatibility by handling problematic characters.
    
    Args:
        unified_data: List of unified log records (from model_dump)
        output_path: Output file path
        version: "original" or "anonymized" to determine which data to use
    """
    if not unified_data: return
    
    def clean_csv_value(value):
        """
        Clean a value for CSV export by handling problematic characters.
        Replaces commas with semicolons and quotes with single quotes.
        """
        if value is None:
            return ""
        value_str = str(value)
        # Replace problematic characters
        value_str = value_str.replace(',', ';')  # Replace commas with semicolons
        value_str = value_str.replace('"', "'")  # Replace double quotes with single quotes
        value_str = value_str.replace('\n', ' ')  # Replace newlines with spaces
        value_str = value_str.replace('\r', ' ')  # Replace carriage returns with spaces
        value_str = value_str.replace('\t', ' ')  # Replace tabs with spaces
        # Trim whitespace
        value_str = value_str.strip()
        return value_str
    
    # Collect all possible keys from parsed_data across all records
    all_keys = set()
    for record in unified_data:
        if record.get('parsed_data'): 
            all_keys.update(record['parsed_data'].keys())
    
    # Sort keys for consistent column order
    sorted_keys = sorted(list(all_keys))
    
    # Define headers with ONLY parsed_data keys (clean and simple)
    headers = sorted_keys
    
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        
        for record in unified_data:
            # Build row with ONLY parsed_data values
            parsed_data = record.get('parsed_data', {})
            row = {}
            
            # Add only parsed_data fields with cleaned values
            for key in sorted_keys:
                row[key] = clean_csv_value(parsed_data.get(key, ""))
            
            writer.writerow(row)

def format_as_excel_analysis(unified_data: List[Dict], output_path: str, config: Dict):
    """
    Generate Excel file with unique values analysis and field insights.
    
    Args:
        unified_data: List of unified log records (from model_dump)
        output_path: Output Excel file path
        config: Application configuration for Presidio and regex patterns
    """
    if not unified_data: return
    
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        from collections import defaultdict, Counter
        import re
    except ImportError as e:
        raise ImportError(f"Required packages not available: {e}. Install with: pip install pandas openpyxl")
    
    # Collect all possible keys and their unique values
    field_values = defaultdict(set)
    field_counts = defaultdict(int)
    
    for record in unified_data:
        parsed_data = record.get('parsed_data', {})
        for key, value in parsed_data.items():
            if value is not None and str(value).strip():
                field_values[key].add(str(value).strip())
                field_counts[key] += 1
    
    # Create workbook
    wb = Workbook()
    
    # Sheet 1: Unique Values Analysis
    ws1 = wb.active
    ws1.title = "Unique Values Analysis"
    
    # Find the maximum number of unique values for any field
    max_unique_values = max(len(values) for values in field_values.values()) if field_values else 0
    
    # Create headers
    headers = list(field_values.keys())
    headers.sort()
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Write unique values
    for col, header in enumerate(headers, 1):
        unique_values = sorted(list(field_values[header]))
        for row, value in enumerate(unique_values, 2):
            cell = ws1.cell(row=row, column=col, value=value)
            # Add alternating row colors for better readability
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 2: Field Analysis Summary
    ws2 = wb.create_sheet("Field Analysis Summary")
    
    # Analysis headers
    summary_headers = [
        "Field Name", "Total Occurrences", "Unique Values", "Most Common Value", 
        "Most Common Count", "Contains IPs", "Contains Emails", "Contains Timestamps",
        "Contains MACs", "Contains URLs", "Contains Hashes", "Contains Ports",
        "Field Type", "Notes"
    ]
    
    # Write summary headers
    for col, header in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Get regex patterns from config
    regex_patterns = config.get('centralized_regex', {})
    detection_patterns = regex_patterns.get('detection', {})
    
    # Compile regex patterns
    compiled_patterns = {}
    for pattern_name, pattern_str in detection_patterns.items():
        try:
            compiled_patterns[pattern_name] = re.compile(pattern_str, re.IGNORECASE)
        except re.error:
            continue
    
    # Analyze each field
    row = 2
    for field_name in sorted(field_values.keys()):
        values = list(field_values[field_name])
        total_occurrences = field_counts[field_name]
        unique_count = len(values)
        
        # Find most common value
        value_counter = Counter()
        for record in unified_data:
            parsed_data = record.get('parsed_data', {})
            if field_name in parsed_data and parsed_data[field_name]:
                value_counter[str(parsed_data[field_name]).strip()] += 1
        
        most_common_value, most_common_count = value_counter.most_common(1)[0] if value_counter else ("", 0)
        
        # Pattern analysis
        contains_ips = any(compiled_patterns.get('ip_address', re.compile(r'')).search(str(v)) for v in values)
        contains_emails = any(compiled_patterns.get('email', re.compile(r'')).search(str(v)) for v in values)
        contains_timestamps = any(compiled_patterns.get('timestamp', re.compile(r'')).search(str(v)) for v in values)
        contains_macs = any(compiled_patterns.get('mac_address', re.compile(r'')).search(str(v)) for v in values)
        contains_urls = any(compiled_patterns.get('url', re.compile(r'')).search(str(v)) for v in values)
        contains_hashes = any(compiled_patterns.get('hash', re.compile(r'')).search(str(v)) for v in values)
        contains_ports = any(compiled_patterns.get('port_number', re.compile(r'')).search(str(v)) for v in values)
        
        # Determine field type
        field_type = "Unknown"
        if contains_ips:
            field_type = "IP Address"
        elif contains_emails:
            field_type = "Email"
        elif contains_timestamps:
            field_type = "Timestamp"
        elif contains_macs:
            field_type = "MAC Address"
        elif contains_urls:
            field_type = "URL"
        elif contains_hashes:
            field_type = "Hash"
        elif contains_ports:
            field_type = "Port"
        elif field_name.lower() in ['timestamp', 'time', 'date']:
            field_type = "Timestamp"
        elif field_name.lower() in ['ip', 'srcip', 'dstip', 'source_ip', 'dest_ip']:
            field_type = "IP Address"
        elif field_name.lower() in ['user', 'username', 'uid']:
            field_type = "User ID"
        elif field_name.lower() in ['port', 'port_number']:
            field_type = "Port"
        
        # Notes
        notes = []
        if contains_ips:
            notes.append("Contains IP addresses")
        if contains_emails:
            notes.append("Contains email addresses")
        if contains_timestamps:
            notes.append("Contains timestamps")
        if contains_macs:
            notes.append("Contains MAC addresses")
        if contains_urls:
            notes.append("Contains URLs")
        if contains_hashes:
            notes.append("Contains hashes")
        if contains_ports:
            notes.append("Contains port numbers")
        
        notes_str = "; ".join(notes) if notes else "Standard field"
        
        # Write row data
        ws2.cell(row=row, column=1, value=field_name)
        ws2.cell(row=row, column=2, value=total_occurrences)
        ws2.cell(row=row, column=3, value=unique_count)
        ws2.cell(row=row, column=4, value=most_common_value)
        ws2.cell(row=row, column=5, value=most_common_count)
        ws2.cell(row=row, column=6, value="Yes" if contains_ips else "No")
        ws2.cell(row=row, column=7, value="Yes" if contains_emails else "No")
        ws2.cell(row=row, column=8, value="Yes" if contains_timestamps else "No")
        ws2.cell(row=row, column=9, value="Yes" if contains_macs else "No")
        ws2.cell(row=row, column=10, value="Yes" if contains_urls else "No")
        ws2.cell(row=row, column=11, value="Yes" if contains_hashes else "No")
        ws2.cell(row=row, column=12, value="Yes" if contains_ports else "No")
        ws2.cell(row=row, column=13, value=field_type)
        ws2.cell(row=row, column=14, value=notes_str)
        
        # Add alternating row colors
        if row % 2 == 0:
            for col in range(1, 15):
                ws2.cell(row=row, column=col).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        row += 1
    
    # Auto-adjust column widths for summary sheet
    for column in ws2.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)  # Cap at 30 characters
        ws2.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_path)

# --- API Endpoints ---

@app.get("/", response_class=HTMLResponse)
async def read_root(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.get("/api/config", response_class=JSONResponse)
async def get_config():
    """
    Returns the current Presidio configuration, augmented with details
    about the recognizers for the UI, by using the PresidioService.
    """
    config_service = ConfigService()
    config = config_service.load_config()
    presidio_config = config.get("presidio", {})

    presidio_service = PresidioService(presidio_config)

    detailed_entities = presidio_service.get_recognizer_details()

    if detailed_entities:
        presidio_config.setdefault("analyzer", {})["entities"] = detailed_entities

    return presidio_config

@app.post("/api/config", response_class=JSONResponse)
async def save_config(update_request: ConfigUpdateRequest):
    config_service = ConfigService()
    full_config = config_service.load_config()
    full_config["presidio"] = update_request.presidio
    if config_service.save_config(full_config):
        return {"status": "success", "message": "Configuration saved successfully."}
    else:
        return JSONResponse(status_code=500, content={"status": "error", "message": "Failed to save configuration."})

@app.get("/api/sample-files", response_class=JSONResponse)
async def get_sample_files():
    examples_dir = "examples"
    try:
        files = [f for f in os.listdir(examples_dir) if os.path.isfile(os.path.join(examples_dir, f))]
        return {"files": files}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/api/sample-line", response_class=JSONResponse)
async def get_sample_line(filepath: str, line_number: int = 1):
    examples_dir = os.path.abspath("examples")
    requested_path = os.path.abspath(os.path.join(examples_dir, os.path.basename(filepath)))
    if not requested_path.startswith(examples_dir):
        return JSONResponse(status_code=403, content={"error": "Access forbidden."})
    try:
        with open(requested_path, 'r', encoding='utf-8', errors='ignore') as f:
            for i, line in enumerate(f):
                if i + 1 == line_number: return {"line_content": line.strip()}
        return JSONResponse(status_code=404, content={"error": f"Line {line_number} not found."})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/api/preview", response_class=JSONResponse)
async def preview_anonymization(preview_request: PreviewRequest):
    presidio_config = preview_request.presidio_config
    sample_text = preview_request.sample_text
    if not sample_text:
        return JSONResponse(content={"anonymized_text": ""})

    try:
        presidio_service = PresidioService(presidio_config)

        if not presidio_service.is_enabled:
            return JSONResponse(content={"anonymized_text": "[PREVIEW] Presidio is disabled."})

        analyzer_config = presidio_config.get("analyzer", {})

        # The entities from the UI can be either a detailed dict or a simple boolean.
        # This logic handles both cases to ensure the preview works correctly.
        entities_map = analyzer_config.get("entities", {})
        enabled_entities = []
        for k, v in entities_map.items():
            if isinstance(v, dict):
                # Handles the detailed structure from the config file
                if v.get('enabled'):
                    enabled_entities.append(k)
            elif v:
                # Handles the simple boolean from the UI's preview request
                enabled_entities.append(k)

        ad_hoc_recognizers = analyzer_config.get('ad_hoc_recognizers', [])
        for rec in ad_hoc_recognizers:
            if rec.get("name") and rec.get("name") not in enabled_entities:
                enabled_entities.append(rec.get("name"))

        if not enabled_entities:
            return JSONResponse(content={"anonymized_text": "[PREVIEW] No entities enabled."})

        anonymized_text = presidio_service.anonymize_text(
            sample_text,
            language=analyzer_config.get("languages", ["en"])[0],
            entities=enabled_entities
        )

        return JSONResponse(content={"anonymized_text": anonymized_text})
    except Exception as e:
        import traceback
        return JSONResponse(status_code=500, content={"error": f"An error occurred during preview: {traceback.format_exc()}"})

@app.post("/api/analysis/{analysis_type}")
async def run_analysis(analysis_type: str, request: AnalysisRequest):
    try:
        config_service = ConfigService()
        config = config_service.load_config()

        log_reader = LogReader(config)
        parser_chain = create_parser_chain(config)
        drain3_service = Drain3Service(config)
        presidio_service = PresidioService(config.get('presidio', {}))

        input_path = os.path.join("examples", request.input_file)
        if not os.path.exists(input_path):
            return JSONResponse(status_code=404, content={"error": "Input file not found."})

        all_records: List[ParsedRecord] = []
        for line_num, line_content in log_reader.read_lines(input_path):
            if not line_content: continue

            log_entry = LogEntry(line_number=line_num, content=line_content, source_file=input_path)
            parsed_record = parser_chain.handle(log_entry)

            if parsed_record:
                parsed_record.presidio_anonymized = presidio_service.anonymize_text(
                    parsed_record.original_content,
                    language=config.get('presidio', {}).get('analyzer', {}).get('languages', ['en'])[0]
                )
                parsed_record.presidio_metadata = []
                all_records.append(parsed_record)

        original_content = [rec.original_content for rec in all_records]
        original_results = drain3_service.process_batch(original_content, 'original')
        anonymized_content = [rec.presidio_anonymized or "" for rec in all_records]
        anonymized_results = drain3_service.process_batch(anonymized_content, 'anonymized')

        for i, record in enumerate(all_records):
            if i < len(original_results): record.drain3_original = original_results[i]
            if i < len(anonymized_results): record.drain3_anonymized = anonymized_results[i]

        # Generate unified data for all analysis types that need it
        unified_data = [record.model_dump(exclude_none=True) for record in all_records]
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename_base = f"{Path(request.input_file).stem}_{analysis_type}_{timestamp}"
        output_filename = ""

        if analysis_type == "anonymize":
            output_filename = f"{output_filename_base}.log"
            output_path = os.path.join("outputs", output_filename)
            format_as_anonymized_text(all_records, output_path)
        elif analysis_type == "logppt":
            # Generate both original and anonymized LogPPT reports
            # Original data report
            original_filename = f"{output_filename_base}_original.csv"
            original_path = os.path.join("outputs", original_filename)
            format_as_logppt(all_records, original_path, "original")
            
            # Anonymized data report
            anonymized_filename = f"{output_filename_base}_anonymized.csv"
            anonymized_path = os.path.join("outputs", anonymized_filename)
            format_as_logppt(all_records, anonymized_path, "anonymized")
            
            # Return both download URLs
            return {
                "original_download_url": f"/outputs/{original_filename}",
                "anonymized_download_url": f"/outputs/{anonymized_filename}",
                "message": "Both original and anonymized LogPPT reports generated successfully"
            }
        elif analysis_type == "json_report":
            output_filename = f"{output_filename_base}.json"
            output_path = os.path.join("outputs", output_filename)
            format_as_json_report(all_records, output_path)
        elif analysis_type == "csv_export":
            # Generate both original and anonymized CSV exports from unified data
            # Original data export
            original_filename = f"{output_filename_base}_original.csv"
            original_path = os.path.join("outputs", original_filename)
            format_as_csv_from_unified(unified_data, original_path, "original")
            
            # Anonymized data export
            anonymized_filename = f"{output_filename_base}_anonymized.csv"
            anonymized_path = os.path.join("outputs", anonymized_filename)
            format_as_csv_from_unified(unified_data, anonymized_path, "anonymized")
            
            # Return both download URLs
            return {
                "original_download_url": f"/outputs/{original_filename}",
                "anonymized_download_url": f"/outputs/{anonymized_filename}",
                "message": "Both original and anonymized CSV exports generated successfully"
            }
        elif analysis_type == "excel_analysis":
            output_filename = f"{output_filename_base}.xlsx"
            output_path = os.path.join("outputs", output_filename)
            format_as_excel_analysis(unified_data, output_path, config)
            return {
                "download_url": f"/outputs/{output_filename}",
                "message": "Excel analysis report generated successfully"
            }
        else:
            return JSONResponse(status_code=400, content={"error": "Invalid analysis type."})

        return {"download_url": f"/outputs/{output_filename}"}
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": f"An error occurred during analysis: {str(e)}"})

@app.post("/api/logppt/run")
async def run_logppt(
    file: UploadFile = File(...),
    model_name: str = Form(...),
    shots: str = Form(...),
    max_train_steps: int = Form(...),
    content_config: str = Form(""),
    columns_order: str = Form("")
):
    try:
        task_id = task_log_manager.start_task()

        # Save the uploaded file
        upload_dir = Path("uploads")
        upload_dir.mkdir(exist_ok=True)
        file_path = upload_dir / file.filename
        with open(file_path, "wb") as buffer:
            buffer.write(await file.read())

        shot_list = [int(s.strip()) for s in shots.split(',')]
        dataset_name = Path(file.filename).stem

        def do_run_pipeline():
            try:
                config_service = ConfigService()
                config = config_service.load_config()
                logppt_service = LogPPTService(config, task_id, task_log_manager)

                results = logppt_service.run_pipeline(
                    file_path=str(file_path),
                    model_name=model_name,
                    shots=shot_list,
                    max_train_steps=max_train_steps,
                    dataset_name=dataset_name,
                    content_config=content_config,
                    columns_order=columns_order
                )
                task_log_manager.log(task_id, f"PIPELINE_COMPLETE::{json.dumps(results)}")
            except Exception as e:
                import traceback
                import sys
                
                # Logging dettagliato dell'errore
                error_msg = f"ERRORE CRITICO nella pipeline LogPPT: {str(e)}"
                task_log_manager.log(task_id, "="*80)
                task_log_manager.log(task_id, error_msg)
                task_log_manager.log(task_id, f"Tipo errore: {type(e).__name__}")
                task_log_manager.log(task_id, f"Modulo: {e.__class__.__module__}")
                
                # Stack trace completo
                full_traceback = traceback.format_exc()
                task_log_manager.log(task_id, "STACK TRACE COMPLETO:")
                for line in full_traceback.split('\n'):
                    if line.strip():
                        task_log_manager.log(task_id, line)
                
                # Informazioni aggiuntive per il debug
                task_log_manager.log(task_id, f"Python version: {sys.version}")
                task_log_manager.log(task_id, f"Working directory: {Path.cwd()}")
                
                # Informazioni sui parametri della chiamata
                task_log_manager.log(task_id, f"Parametri chiamata:")
                task_log_manager.log(task_id, f"  - model_name: {model_name}")
                task_log_manager.log(task_id, f"  - shots: {shot_list}")
                task_log_manager.log(task_id, f"  - max_train_steps: {max_train_steps}")
                task_log_manager.log(task_id, f"  - dataset_name: {dataset_name}")
                task_log_manager.log(task_id, f"  - file_path: {file_path}")
                
                task_log_manager.log(task_id, "="*80)
                task_log_manager.log(task_id, f"ERROR::{str(e)}")
            finally:
                task_log_manager.finish_task(task_id)

        thread = Thread(target=do_run_pipeline)
        thread.start()

        return {"task_id": task_id}

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": f"An error occurred during LogPPT processing: {str(e)}"})


@app.post("/api/logppt2/run")
async def run_logppt2(
    file: UploadFile | None = File(None),
    shots: str = Form("8, 16"),
    evaluate: bool = Form(False),
    dataset_filename: str = Form("")
):
    try:
        task_id = task_log_manager.start_task()

        file_path: Path | None = None

        if dataset_filename:
            # Usa file dal dataset_dir configurato
            config_service = ConfigService()
            config = config_service.load_config()
            dataset_dir = Path(config.get('logppt2', {}).get('dataset_dir', 'examples'))
            candidate = (dataset_dir / Path(dataset_filename).name)
            if not candidate.exists():
                return JSONResponse(status_code=404, content={"error": f"File non trovato in dataset_dir: {candidate}"})
            file_path = candidate
        elif file is not None:
            # Salva file caricato in uploads
            upload_dir = Path("uploads")
            upload_dir.mkdir(exist_ok=True)
            file_path = upload_dir / file.filename
            with open(file_path, "wb") as buffer:
                buffer.write(await file.read())
        else:
            return JSONResponse(status_code=400, content={"error": "Nessun file fornito: seleziona un file dal dataset configurato o carica un CSV."})

        # Parse shots
        try:
            shot_list = [int(s.strip()) for s in str(shots).split(',') if s.strip()]
        except Exception:
            return JSONResponse(status_code=400, content={"error": "Parametro 'shots' non valido. Usa formato: 8,16,32"})

        def do_run_pipeline():
            try:
                config_service = ConfigService()
                config = config_service.load_config()
                service = LogPPT2Service(config, task_id, task_log_manager)

                results = service.run_pipeline(
                    dataset_path=str(file_path),
                    shots=shot_list,
                    evaluate=bool(evaluate),
                )
                task_log_manager.log(task_id, f"PIPELINE_COMPLETE::{json.dumps(results)}")
            except Exception as e:
                import traceback
                err = traceback.format_exc()
                task_log_manager.log(task_id, err)
                task_log_manager.log(task_id, f"ERROR::{str(e)}")
            finally:
                task_log_manager.finish_task(task_id)

        thread = Thread(target=do_run_pipeline)
        thread.start()

        return {"task_id": task_id}

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": f"An error occurred during LogPPT 2 processing: {str(e)}"})
